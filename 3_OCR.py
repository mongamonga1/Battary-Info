# -*- coding: utf-8 -*-
"""
Streamlit 앱: 사업자등록증 OCR & 국세청 진위확인
"""

import time
import re
import unicodedata
from io import BytesIO

import numpy as np
import pandas as pd
import requests
import streamlit as st
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import Font

# ------------------------------  ⚙️  기본 설정  ------------------------------
st.set_page_config(page_title="사업자등록증 OCR & 진위확인", layout="centered")

# ★★ [추가] EasyOCR Reader를 세션에 프리로드 ──────────────────────────
if "ocr_reader" not in st.session_state:
    with st.spinner("🔄 EasyOCR 모델을 불러오는 중입니다... (최초 10-20초 소요)"):
        st.session_state["ocr_reader"] = get_reader()
# ────────────────────────────────────────────────────────────────────────────

LANGS = ["ko", "en"]  # EasyOCR 언어 설정
DATE_REGEX = re.compile(r"(\d{4})[.\-년 ]+(\d{1,2})[.\-월 ]+(\d{1,2})")
B_NO_REGEX = re.compile(r"\d{3}-\d{2}-\d{5}|\d{10}")
SERVICE_KEY = st.secrets.get("ODCLOUD_SERVICE_KEY", "")  # 국세청 일반 인증키

# ------------------------------
# 📚  헬퍼 함수
# ------------------------------
@st.cache_resource(show_spinner=False)
def get_reader():
    """EasyOCR Reader는 초기화 비용이 크므로 캐싱"""
    import easyocr  # 지연 임포트

    return easyocr.Reader(LANGS)

def normalize_text(s: str) -> str:
    """텍스트 정규화"""
    if s is None:
        return ""
    s = str(s).strip()
    s = re.sub(r"[\u200B-\u200D\uFEFF]", "", s)  # Zero‑width 제거
    s = re.sub(r"\s+", " ", s)  # 공백 정리
    return unicodedata.normalize("NFC", s)

def extract_info(lines):
    """
    OCR 라인에서
      · b_no      : 10자리 사업자번호
      · start_dt  : 개업연월일(YYYYMMDD)
      · p_nm      : 대표자/성명
    추출 (라벨·이름 사이 공백 뒤틀림 대응)
    """
    # ── 전처리 ───────────────────────────────────
    # ① 원본 라인 그대로(공백 유지)  ② 모든 공백 제거 버전
    joined        = normalize_text(" ".join(lines))
    joined_nospace = re.sub(r"\s+", "", joined)

    info = {"b_no": None, "start_dt": None, "p_nm": None}

    # ── 1) 사업자등록번호 ────────────────────────
    m = re.search(r"\d{3}-\d{2}-\d{5}|\d{10}", joined_nospace)
    if m:
        info["b_no"] = m.group().replace("-", "")

    # ── 2) 대표자 / 성명 (공백·콜론·한자점 모두 허용) ────────
    m = re.search(r"(대표자|성명)(:|：)?([가-힣]{2,10})", joined_nospace)
    if m:
        info["p_nm"] = m.group(3)

    # ── 3) 개업 연월일 ───────────────────────────
    #    (라벨이 있을 때 우선, 없으면 날짜 패턴만)
    m = re.search(
        r"개업[^0-9]{0,10}"
        r"(\d{4})[.\-년 ]+(\d{1,2})[.\-월 ]+(\d{1,2})",
        joined,
    )
    if not m:
        m = re.search(r"(\d{4})[.\-년 ]+(\d{1,2})[.\-월 ]+(\d{1,2})", joined)
    if m:
        y, mth, d = m.groups()
        info["start_dt"] = f"{y}{int(mth):02d}{int(d):02d}"

    return info

def ocr_image(file) -> list[str]:
    """업로드된 이미지 파일에서 OCR 수행 후 텍스트 라인 반환"""
    img = Image.open(file)
    reader = st.session_state["ocr_reader"]
    result = reader.readtext(np.array(img))
    return [text for _, text, _ in result]


def validate_business(df: pd.DataFrame) -> pd.DataFrame:
    """국세청 API로 진위확인"""
    if not SERVICE_KEY:
        st.warning("❗️ 먼저 Settings → Secrets 에 `ODCLOUD_SERVICE_KEY`를 등록하세요.")
        return df

    url = f"https://api.odcloud.kr/api/nts-businessman/v1/validate?serviceKey={SERVICE_KEY}"
    headers = {"Content-Type": "application/json", "Accept": "application/json"}

    valid_list, msg_list = [], []
    for _, row in df.iterrows():
        payload = {
            "businesses": [
                {
                    "b_no": str(row["b_no"]).zfill(10),
                    "start_dt": str(row["start_dt"]),
                    "p_nm": str(row["p_nm"]),
                    "p_nm2": "",
                    "b_nm": "",
                    "corp_no": "",
                    "b_sector": "",
                    "b_type": "",
                    "b_adr": "",
                }
            ]
        }
        try:
            resp = requests.post(url, headers=headers, json=payload, timeout=10)
            data = resp.json()["data"][0]
            code = data["valid"]
            msg = data.get("valid_msg", "")
            res = "일치" if code == "01" else "불일치" if code == "02" else f"기타({code})"
        except Exception as e:
            res, msg = "에러", str(e)
        valid_list.append(res)
        msg_list.append(msg)
        time.sleep(0.2)  # 🎗 API 과다 호출 방지

    df = df.copy()
    df["진위확인결과"] = valid_list
    df["API메시지"] = msg_list
    return df


def to_colored_excel(df: pd.DataFrame) -> bytes:
    """결과 컬럼 색상 강조 후 바이너리 반환"""
    bio = BytesIO()
    df.to_excel(bio, index=False)
    bio.seek(0)
    wb = load_workbook(bio)
    ws = wb.active
    col = df.columns.get_loc("진위확인결과") + 1
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        cell = row[col - 1]
        if cell.value == "일치":
            cell.font = Font(color="008000")
        elif cell.value == "불일치":
            cell.font = Font(color="FF0000")
        elif cell.value == "에러":
            cell.font = Font(color="808080")
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()

# ------------------------------
# 🖼  UI 구성
# ------------------------------
st.title("📑 사업자등록증 OCR & 진위확인")

img_tab, file_tab = st.tabs(["이미지 업로드", "CSV/XLSX 업로드"])

with img_tab:
    images = st.file_uploader(
        "✅ 사업자등록증 이미지를 선택하세요 (다중 선택 가능)",
        type=["png", "jpg", "jpeg"],
        accept_multiple_files=True,
    )
    # ① -------------- 디버깅 토글 추가 --------------
    debug = st.checkbox("🔍 OCR 원본 라인 보기", value=False)
    # ------------------------------------------------
    if images:
        rows = []
        with st.spinner("🔍 OCR 분석 중..."):
            for f in images:
                lines = ocr_image(f)
                rows.append(extract_info(lines))
        df_img = pd.DataFrame(rows).dropna(how="all")
        st.subheader("추출된 정보")
        st.dataframe(df_img, use_container_width=True)

        if st.button("국세청 API로 진위확인"):
            df_valid = validate_business(df_img)
            st.dataframe(df_valid, use_container_width=True)
            excel_bytes = to_colored_excel(df_valid)
            st.download_button(
                "📥 결과 Excel 다운로드",
                data=excel_bytes,
                file_name="validation_result.xlsx",
            )

with file_tab:
    data_file = st.file_uploader("✅ 사업자 정보가 담긴 CSV 또는 XLSX", type=["csv", "xlsx"])
    # ③ -------------- 디버깅 토글 추가 --------------
    debug2 = st.checkbox("🔍 OCR 원본(샘플) 보기", value=False)
    # ------------------------------------------------
    if data_file is not None:
        df = (
            pd.read_csv(data_file) if data_file.name.endswith(".csv") else pd.read_excel(data_file)
        )
        st.dataframe(df, use_container_width=True)

        if st.button("API 진위확인 실행"):
            df_valid = validate_business(df)
            st.dataframe(df_valid, use_container_width=True)
            excel_bytes = to_colored_excel(df_valid)
            st.download_button(
                "📥 결과 Excel 다운로드",
                data=excel_bytes,
                file_name="validation_result.xlsx",
            )

st.markdown("---")
st.info(
    "- ⏱ EasyOCR 최초 실행은 수 초가 걸릴 수 있습니다.\n"
    "- 💡 국세청 API는 초당 5회 이하로 호출하세요."
)
